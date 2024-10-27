import json
import os
import pickle

import numpy as np
import pandas as pd
import torch
import yaml
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import Alignment
from transformers import AutoModelForCausalLM, AutoTokenizer
from tqdm import tqdm
import utils
from easyeditor import BaseEditor
from easyeditor import ROMEHyperParams
from easyeditor.models.rome.causal_trace import guess_subject
from easyeditor.util import nethook

benchmark = pd.read_excel("./benchmark/FactEditing.xlsx", sheet_name="Tabelle1")
prompts = list(benchmark.Prompt)

hparams_path = "./hparams/ROME/gpt2-xl.yaml"
config = yaml.safe_load(open(hparams_path, "r"))

tokenizer = AutoTokenizer.from_pretrained(config["model_name"])
tokenizer.pad_token_id = tokenizer.eos_token_id
tokenizer.padding_side='left'

model = AutoModelForCausalLM.from_pretrained(config["model_name"]).to('cuda')

def generate_outputs(generation_prompts, generation_model):
    batch = tokenizer(generation_prompts, return_tensors='pt', padding=True)

    outputs = generation_model.generate(
        input_ids=batch['input_ids'].to('cuda'),
        attention_mask=batch['attention_mask'].to('cuda'),
        max_new_tokens=50,
        do_sample=True,
        top_k=0,
        temperature=.001
    )

    return [tokenizer.decode(x).replace("<|end_of_text|>", "").replace("<|endoftext|>", "").replace("<|begin_of_text|>", "").replace("\xa0", "") for x in outputs.detach().cpu().numpy().tolist()]

benchmark["Pre-Edit"] = generate_outputs(prompts, model)

del model

rome_test = utils.RomeTest(hparams_path)

def causal_tracing_test(tracing_prompt, subject):
    print(f"Causal Tracing for: {tracing_prompt}")
    max_effect_layer, min_effect_layer, hidden_flow = rome_test.get_hidden_flow(tracing_prompt, subject)
    print(f"Max Effect Layer (last subject token): {max_effect_layer}")
    print(f"Min Effect Layer (last subject token): {min_effect_layer}")
    rome_test.plot_causal_tracing_effect(tracing_prompt, kind="mlp", hidden_flow=hidden_flow)
    return max_effect_layer, min_effect_layer


max_effect_layers = dict()
min_effect_layers = dict()
for i, prompt in tqdm(enumerate(prompts)):
    max_effect_layers[prompt], min_effect_layers[prompt] = causal_tracing_test(prompt, benchmark.iloc[i]["Subject"])

del rome_test

def edit_model(dataframe):
    filename = f"./layer_store/{config['model_name'].replace('/', '-')}/layer-{config['layers'][0]}-prompt{min(dataframe.index)}-{max(dataframe.index)}.weight"
    parameter_name = f'{config["rewrite_module_tmp"].format(config["layers"][0])}.weight'
    try:
        weights = pickle.load(open(filename, "rb"))
        edited_model = AutoModelForCausalLM.from_pretrained(config["model_name"]).to('cuda')
        with torch.no_grad():
            w = nethook.get_parameter(edited_model, parameter_name)
            w[...] = weights
    except FileNotFoundError:
        hparams = ROMEHyperParams.from_hparams(hparams_path)
        editor = BaseEditor.from_hparams(hparams)
        _, edited_model, _ = editor.edit(
            prompts=list(dataframe["Prompt"]),
            ground_truth=list(dataframe["Ground Truth"]),
            target_new=list(dataframe["Target"]),
            subject=list(dataframe["Subject"]),
            keep_original_weight=False
        )
        os.makedirs(os.path.dirname(filename), exist_ok=True)
        pickle.dump(nethook.get_parameter(edited_model, parameter_name), open(filename, "wb"))
    return edited_model

def benchmark_rome(dataframe, option):
    original_layer = config["layers"][0]
    if option == "max":
        # noinspection PyTypeChecker
        edit_layer = round(np.median([max_effect_layers[x] for x in list(dataframe["Prompt"])]))
    elif option == "min":
        # noinspection PyTypeChecker
        edit_layer = round(np.median([min_effect_layers[x] for x in list(dataframe["Prompt"])]))
    elif option == "last":
        edit_layer = config["v_loss_layer"]
    else:
        edit_layer = original_layer
    config["layers"][0] = edit_layer
    yaml.safe_dump(config, open(hparams_path, "w"))
    edited_model = edit_model(dataframe)
    config["layers"][0] = original_layer
    yaml.safe_dump(config, open(hparams_path, "w"))
    benchmark["Post-Edit"] = generate_outputs(prompts, edited_model)
    filename = f"./benchmark/results/{config['model_name'].replace('/', '-')}-{option}-layer-prompt-{min(dataframe.index)}-{max(dataframe.index)}.xlsx"
    os.makedirs(os.path.dirname(filename), exist_ok=True)
    benchmark.to_excel(filename, index=None)
    wb = load_workbook(filename)
    ws = wb.active

    # Spaltenbreiten setzen
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 50
    ws.column_dimensions['F'].width = 50

    # Zeilenumbruch für die Zellen in Spalten A bis E aktivieren
    for row in ws.iter_rows(min_col=1, max_col=6):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True)

    # Datei speichern
    wb.save(filename)


def benchmark_all_options(dataframe):
    for option in ["max", "min", "last", "recom"]:
        benchmark_rome(dataframe, option)

torch.set_grad_enabled(True)

# Single Edit
for i in range(len(benchmark.index)):
    dataframe = benchmark.iloc[[i]]
    benchmark_all_options(dataframe)

# Multi Edit
benchmark_all_options(benchmark)